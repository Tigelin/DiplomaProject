from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from journal.models import Grade, Task, Discipline, Lesson, LessonFile, Attendance, Group, Student, Schedule, LessonType, AttendanceType
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from django.http import HttpResponse
from urllib.parse import quote
from datetime import datetime

# Create your views here.


def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            messages.success(request, f'Добро пожаловать, {user.get_full_name() or user.username}!')
            return redirect('home')
        else:
            messages.error(request, 'Неверное имя пользователя или пароль.')

    return render(request, 'users/login.html')


def logout_view(request):
    logout(request)
    messages.success(request, 'Вы успешно вышли из системы.')
    return redirect('home')


@login_required
def dashboard(request):
    if request.user.role and request.user.role.name == 'Студент':
        return redirect('student_dashboard')
    elif request.user.role and request.user.role.name == 'Преподаватель':
        return redirect('teacher_dashboard')
    return render(request, 'users/dashboard.html')


@login_required
def student_dashboard(request):
    try:
        student = request.user.student
    except:
        messages.error(request, 'Профиль студента не найден.')
        return redirect('home')

    return render(request, 'users/student/dashboard.html', {'student': student})


@login_required
def student_profile(request):
    try:
        student = request.user.student
    except:
        messages.error(request, 'Профиль студента не найден.')
        return redirect('home')

    if request.method == 'POST':
        user = request.user
        user.last_name = request.POST.get('last_name')
        user.first_name = request.POST.get('first_name')
        user.patronymic = request.POST.get('patronymic')
        user.phone = request.POST.get('phone')
        user.save()

        student.special_number = request.POST.get('special_number')
        student.save()

        messages.success(request, 'Профиль успешно обновлён!')
        return redirect('student_profile')

    return render(request, 'users/student/profile.html', {
        'student': student,
    })


@login_required
def student_grades(request):
    try:
        student = request.user.student
    except:
        messages.error(request, 'Профиль студента не найден.')
        return redirect('home')

    grades = Grade.objects.filter(
        student=student,
        task__isnull=False,
        task__lesson__isnull=False
    ).select_related(
        'task__lesson__schedule__discipline__plan',
        'task__lesson__schedule'
    ).order_by('task__lesson__schedule__date')

    disciplines_dict = {}
    for grade in grades:
        discipline_name = grade.task.lesson.schedule.discipline.plan.name
        discipline_id = grade.task.lesson.schedule.discipline.id
        if discipline_id not in disciplines_dict:
            disciplines_dict[discipline_id] = discipline_name

    dates_dict = {}
    for grade in grades:
        date = grade.task.lesson.schedule.date
        if date not in dates_dict:
            dates_dict[date] = date

    sorted_dates = sorted(dates_dict.keys())

    matrix = {}
    for discipline_id in disciplines_dict:
        matrix[discipline_id] = {}
        for date in sorted_dates:
            matrix[discipline_id][date] = []

    for grade in grades:
        discipline_id = grade.task.lesson.schedule.discipline.id
        date = grade.task.lesson.schedule.date
        matrix[discipline_id][date].append(grade)

    averages = {}
    for discipline_id in disciplines_dict:
        all_grades = []
        for date in sorted_dates:
            for grade in matrix[discipline_id][date]:
                if grade.value == 1:
                    all_grades.append(2)
                else:
                    all_grades.append(grade.value)
        if all_grades:
            avg = sum(all_grades) / len(all_grades)
            averages[discipline_id] = round(avg, 2)
        else:
            averages[discipline_id] = None

    context = {
        'student': student,
        'disciplines': disciplines_dict.items(),
        'dates': sorted_dates,
        'matrix': matrix,
        'averages': averages,
    }
    return render(request, 'users/student/grades.html', context)


@login_required
def student_tasks(request):
    try:
        student = request.user.student
    except:
        messages.error(request, 'Профиль студента не найден.')
        return redirect('home')

    show_all = request.GET.get('show_all', 'false') == 'true'
    discipline_id = request.GET.get('discipline_id', '')

    grades_dict = {}
    for grade in Grade.objects.filter(student=student, task__isnull=False):
        grades_dict[grade.task_id] = grade.value

    completed_task_ids = [task_id for task_id, grade in grades_dict.items() if grade >= 2]

    all_tasks = Task.objects.filter(
        lesson__schedule__discipline__group=student.group
    ).select_related(
        'lesson__schedule__discipline__plan',
        'lesson__schedule__discipline__teacher__user'
    ).distinct().order_by('lesson__schedule__date')

    if discipline_id:
        all_tasks = all_tasks.filter(lesson__schedule__discipline__id=discipline_id)

    tasks_with_status = []
    for task in all_tasks:
        if task.id in grades_dict:
            grade = grades_dict[task.id]
            is_completed = grade >= 2
            tasks_with_status.append({
                'task': task,
                'is_completed': is_completed,
                'grade': grade,
            })

    if not show_all:
        tasks_with_status = [t for t in tasks_with_status if not t['is_completed']]

    disciplines = Discipline.objects.filter(group=student.group).select_related('plan')

    context = {
        'student': student,
        'tasks_with_status': tasks_with_status,
        'disciplines': disciplines,
        'show_all': show_all,
        'selected_discipline_id': discipline_id,
    }
    return render(request, 'users/student/tasks.html', context)


@login_required
def lesson_detail(request, lesson_id):
    try:
        student = request.user.student
    except:
        messages.error(request, 'Профиль студента не найден.')
        return redirect('home')

    lesson = get_object_or_404(Lesson, id=lesson_id)

    if lesson.schedule.discipline.group != student.group:
        messages.error(request, 'У вас нет доступа к этому занятию.')
        return redirect('student_grades')

    tasks = Task.objects.filter(lesson=lesson)

    tasks_with_grades = []
    for task in tasks:
        grade = Grade.objects.filter(student=student, task=task).first()
        if grade:
            tasks_with_grades.append({
                'task': task,
                'grade': grade.value,
            })

    files = LessonFile.objects.filter(lesson=lesson)

    context = {
        'lesson': lesson,
        'tasks_with_grades': tasks_with_grades,
        'files': files,
    }
    return render(request, 'users/student/lesson_detail.html', context)


@login_required
def student_attendance(request):
    try:
        student = request.user.student
    except:
        messages.error(request, 'Профиль студента не найден.')
        return redirect('home')

    attendances = Attendance.objects.filter(student=student).select_related(
        'lesson__schedule__discipline__plan',
        'lesson__schedule',
        'attendance_type'
    ).order_by('lesson__schedule__date')

    disciplines_dict = {}
    for att in attendances:
        discipline_name = att.lesson.schedule.discipline.plan.name
        discipline_id = att.lesson.schedule.discipline.id
        if discipline_id not in disciplines_dict:
            disciplines_dict[discipline_id] = discipline_name

    dates_dict = {}
    for att in attendances:
        date = att.lesson.schedule.date
        if date not in dates_dict:
            dates_dict[date] = date

    sorted_dates = sorted(dates_dict.keys())

    matrix = {}
    for discipline_id in disciplines_dict:
        matrix[discipline_id] = {}
        for date in sorted_dates:
            matrix[discipline_id][date] = None

    for att in attendances:
        discipline_id = att.lesson.schedule.discipline.id
        date = att.lesson.schedule.date
        matrix[discipline_id][date] = att.attendance_type.name

    total = attendances.count()
    present = attendances.filter(attendance_type__name='Присутствовал').count()
    absent = total - present
    attendance_percent = round((present / total * 100) if total > 0 else 0)

    context = {
        'student': student,
        'disciplines': disciplines_dict.items(),
        'dates': sorted_dates,
        'matrix': matrix,
        'total': total,
        'present': present,
        'absent': absent,
        'attendance_percent': attendance_percent,
    }
    return render(request, 'users/student/attendance.html', context)


@login_required
def teacher_dashboard(request):
    try:
        teacher = request.user.teacher
    except:
        messages.error(request, 'Профиль преподавателя не найден.')
        return redirect('home')

    context = {
        'teacher': teacher,
    }
    return render(request, 'users/teacher/dashboard.html', context)


@login_required
def teacher_profile(request):
    try:
        teacher = request.user.teacher
    except:
        messages.error(request, 'Профиль преподавателя не найден.')
        return redirect('home')

    context = {
        'teacher': teacher,
    }
    return render(request, 'users/teacher/profile.html', context)


@login_required
def teacher_groups(request):
    try:
        teacher = request.user.teacher
    except:
        messages.error(request, 'Профиль преподавателя не найден.')
        return redirect('home')

    groups = Group.objects.filter(
        discipline__teacher=teacher
    ).distinct()

    for group in groups:
        disciplines = Discipline.objects.filter(group=group, teacher=teacher)
        group.disciplines = disciplines
        for discipline in disciplines:
            total_hours = 0
            lessons = Lesson.objects.filter(
                schedule__discipline=discipline
            )
            for lesson in lessons:
                total_hours += lesson.hours
            discipline.actual_hours = total_hours

    context = {
        'teacher': teacher,
        'groups': groups,
    }
    return render(request, 'users/teacher/groups.html', context)


@login_required
def teacher_journal(request, discipline_id):
    try:
        teacher = request.user.teacher
    except:
        messages.error(request, 'Профиль преподавателя не найден.')
        return redirect('home')

    discipline = get_object_or_404(Discipline, id=discipline_id)

    if discipline.teacher != teacher:
        messages.error(request, 'У вас нет доступа к этой дисциплине.')
        return redirect('teacher_groups')

    students = Student.objects.filter(group=discipline.group).select_related('user').order_by('user__last_name')

    schedules = Schedule.objects.filter(
        discipline=discipline
    ).select_related('classroom').order_by('date', 'lesson_number')

    per_page = request.GET.get('per_page', 5)
    try:
        per_page = int(per_page)
    except:
        per_page = 5
    if per_page not in [5, 10, 20, 50]:
        per_page = 5

    paginator = Paginator(schedules, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    for schedule in page_obj:
        lesson = Lesson.objects.filter(schedule=schedule).first()
        schedule.has_lesson = lesson is not None
        if schedule.has_lesson:
            schedule.lesson = lesson
            schedule.lesson_type = lesson.lesson_type.name if lesson.lesson_type else '—'
            schedule.tasks = Task.objects.filter(lesson=lesson)
        else:
            schedule.lesson_type = None
            schedule.tasks = []
            schedule.lesson = None

    grades_matrix = {}
    for student in students:
        grades_matrix[student.id] = {}
        for schedule in page_obj:
            if schedule.has_lesson:
                for task in schedule.tasks:
                    grades_matrix[student.id][task.id] = None

    grades = Grade.objects.filter(
        task__lesson__schedule__discipline=discipline
    ).select_related('student', 'task')

    for grade in grades:
        student_id = grade.student.id
        task_id = grade.task.id
        if student_id in grades_matrix and task_id in grades_matrix[student_id]:
            grades_matrix[student_id][task_id] = grade.value

    context = {
        'teacher': teacher,
        'discipline': discipline,
        'students': students,
        'schedules': page_obj,
        'grades_matrix': grades_matrix,
        'per_page': per_page,
    }
    return render(request, 'users/teacher/journal.html', context)


@login_required
def teacher_lesson(request, schedule_id):
    try:
        teacher = request.user.teacher
    except:
        messages.error(request, 'Профиль преподавателя не найден.')
        return redirect('home')

    schedule = get_object_or_404(Schedule, id=schedule_id)

    if schedule.discipline.teacher != teacher:
        messages.error(request, 'У вас нет доступа к этому занятию.')
        return redirect('teacher_groups')

    lesson = Lesson.objects.filter(schedule=schedule).first()

    if request.method == 'POST':
        if 'save_lesson' in request.POST:
            if not lesson:
                lesson = Lesson.objects.create(schedule=schedule, hours=2)
                students = Student.objects.filter(group=schedule.discipline.group)
                attendance_type, _ = AttendanceType.objects.get_or_create(name='Присутствовал')
                for student in students:
                    Attendance.objects.create(lesson=lesson, student=student, attendance_type=attendance_type)
            else:
                lesson.hours = 2
                lesson.save()

            lesson.topic = request.POST.get('topic', '')
            lesson_type_id = request.POST.get('lesson_type')
            if lesson_type_id:
                lesson.lesson_type_id = lesson_type_id
            lesson.save()

            messages.success(request, 'Занятие сохранено.')
            return redirect('teacher_lesson', schedule_id=schedule.id)

        elif 'add_task' in request.POST:
            Task.objects.create(
                name=request.POST.get('task_name'),
                lesson=lesson,
                description=request.POST.get('task_description', '')
            )
            messages.success(request, 'Задание добавлено.')
            return redirect('teacher_lesson', schedule_id=schedule.id)

        elif 'upload_file' in request.POST:
            if request.FILES.get('file'):
                LessonFile.objects.create(
                    name=request.POST.get('file_name', request.FILES['file'].name),
                    lesson=lesson,
                    file=request.FILES['file']
                )
                messages.success(request, 'Файл загружен.')
            return redirect('teacher_lesson', schedule_id=schedule.id)

    created = lesson is None

    lesson_types = LessonType.objects.all()
    files = LessonFile.objects.filter(lesson=lesson) if lesson else []
    tasks = Task.objects.filter(lesson=lesson) if lesson else []

    context = {
        'teacher': teacher,
        'schedule': schedule,
        'lesson': lesson,
        'created': created,
        'lesson_types': lesson_types,
        'files': files,
        'tasks': tasks,
    }
    return render(request, 'users/teacher/lesson.html', context)


@login_required
def teacher_task_grades(request, task_id=None, lesson_id=None):
    try:
        teacher = request.user.teacher
    except:
        messages.error(request, 'Профиль преподавателя не найден.')
        return redirect('home')

    if task_id:
        task = get_object_or_404(Task, id=task_id)
        lesson = task.lesson
        schedule = lesson.schedule
    else:
        lesson = get_object_or_404(Lesson, id=lesson_id)
        schedule = lesson.schedule
        task, created = Task.objects.get_or_create(
            lesson=lesson,
            name='Новое задание'
        )

    if schedule.discipline.teacher != teacher:
        messages.error(request, 'У вас нет доступа.')
        return redirect('teacher_groups')

    students = Student.objects.filter(group=schedule.discipline.group).select_related('user').order_by(
        'user__last_name')

    if request.method == 'POST':
        if 'delete' in request.POST:
            task.delete()
            messages.success(request, 'Задание удалено.')
            return redirect('teacher_journal', discipline_id=schedule.discipline.id)

        else:
            task.name = request.POST.get('task_name')
            task.description = request.POST.get('task_description')
            task.save()

            for student in students:
                grade_value = request.POST.get(f'grade_{student.id}')
                if grade_value and grade_value.isdigit() and 1 <= int(grade_value) <= 5:
                    Grade.objects.update_or_create(
                        task=task,
                        student=student,
                        defaults={'value': int(grade_value)}
                    )
                elif grade_value == '':
                    Grade.objects.filter(task=task, student=student).delete()

            messages.success(request, 'Задание и оценки сохранены.')
            return redirect('teacher_task_grades', task_id=task.id)

    grades = {}
    for student in students:
        grade = Grade.objects.filter(task=task, student=student).first()
        grades[student.id] = grade.value if grade else None

    context = {
        'teacher': teacher,
        'task': task,
        'schedule': schedule,
        'lesson': lesson,
        'students': students,
        'grades': grades,
    }
    return render(request, 'users/teacher/task_grades.html', context)


@login_required
def teacher_task_create(request, lesson_id):
    try:
        teacher = request.user.teacher
    except:
        messages.error(request, 'Профиль преподавателя не найден.')
        return redirect('home')

    lesson = get_object_or_404(Lesson, id=lesson_id)

    if lesson.schedule.discipline.teacher != teacher:
        messages.error(request, 'У вас нет доступа.')
        return redirect('teacher_groups')

    task = Task.objects.create(lesson=lesson, name='Новое задание')

    return redirect('teacher_task_grades', task_id=task.id)


@login_required
def export_journal_excel(request, discipline_id):
    try:
        teacher = request.user.teacher
    except:
        messages.error(request, 'Профиль преподавателя не найден.')
        return redirect('home')

    discipline = get_object_or_404(Discipline, id=discipline_id)

    if discipline.teacher != teacher:
        messages.error(request, 'У вас нет доступа к этой дисциплине.')
        return redirect('teacher_groups')

    students = Student.objects.filter(group=discipline.group).select_related('user').order_by('user__last_name')

    schedules = Schedule.objects.filter(
        discipline=discipline
    ).select_related('classroom').order_by('date', 'lesson_number')

    for schedule in schedules:
        lesson = Lesson.objects.filter(schedule=schedule).first()
        schedule.has_lesson = lesson is not None
        if schedule.has_lesson:
            schedule.lesson = lesson
            schedule.lesson_type = lesson.lesson_type.name if lesson.lesson_type else '—'
            schedule.tasks = Task.objects.filter(lesson=lesson)
            schedule.topic = lesson.topic if lesson.topic else '—'
        else:
            schedule.lesson_type = None
            schedule.tasks = []
            schedule.lesson = None
            schedule.topic = None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Журнал_{discipline.plan.name}"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_font = Font(color="ffffff", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    row = 1
    col = 1

    ws.cell(row=row, column=col).value = "Журнал оценок"
    ws.cell(row=row, column=col).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1 + len(schedules))
    row += 2

    ws.cell(row=row, column=col).value = f"Дисциплина: {discipline.plan.name}"
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1 + len(schedules))
    row += 1

    ws.cell(row=row, column=col).value = f"Группа: {discipline.group.name}"
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1 + len(schedules))
    row += 1

    ws.cell(row=row,
            column=col).value = f"Преподаватель: {teacher.user.last_name} {teacher.user.first_name} {teacher.user.patronymic}"
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1 + len(schedules))
    row += 2

    first_header_row = row
    col = 1

    ws.cell(row=first_header_row, column=col).value = "№"
    ws.cell(row=first_header_row, column=col).fill = header_fill
    ws.cell(row=first_header_row, column=col).font = header_font
    ws.cell(row=first_header_row, column=col).alignment = center_alignment
    ws.cell(row=first_header_row, column=col).border = thin_border
    col += 1

    ws.cell(row=first_header_row, column=col).value = "Студент"
    ws.cell(row=first_header_row, column=col).fill = header_fill
    ws.cell(row=first_header_row, column=col).font = header_font
    ws.cell(row=first_header_row, column=col).alignment = center_alignment
    ws.cell(row=first_header_row, column=col).border = thin_border
    col += 1

    for schedule in schedules:
        if schedule.has_lesson and schedule.tasks:
            first_col = col
            for task in schedule.tasks:
                col += 1
            last_col = col - 1
            cell_value = f"{schedule.date.strftime('%d.%m.%Y')} ({schedule.lesson_number} пара)\n{schedule.topic}"
            ws.cell(row=first_header_row, column=first_col).value = cell_value
            ws.cell(row=first_header_row, column=first_col).fill = header_fill
            ws.cell(row=first_header_row, column=first_col).font = header_font
            ws.cell(row=first_header_row, column=first_col).alignment = center_alignment
            ws.cell(row=first_header_row, column=first_col).border = thin_border
            if first_col != last_col:
                ws.merge_cells(start_row=first_header_row, start_column=first_col, end_row=first_header_row,
                               end_column=last_col)
        elif schedule.has_lesson:
            cell_value = f"{schedule.date.strftime('%d.%m.%Y')} ({schedule.lesson_number} пара)\n{schedule.topic}"
            ws.cell(row=first_header_row, column=col).value = cell_value
            ws.cell(row=first_header_row, column=col).fill = header_fill
            ws.cell(row=first_header_row, column=col).font = header_font
            ws.cell(row=first_header_row, column=col).alignment = center_alignment
            ws.cell(row=first_header_row, column=col).border = thin_border
            col += 1

    second_header_row = first_header_row + 1
    col = 1

    for i in range(1, 3):
        ws.cell(row=second_header_row, column=i).value = ""
        ws.cell(row=second_header_row, column=i).fill = header_fill
        ws.cell(row=second_header_row, column=i).border = thin_border
        if i == 1:
            ws.merge_cells(start_row=first_header_row, start_column=1, end_row=second_header_row, end_column=1)
        elif i == 2:
            ws.merge_cells(start_row=first_header_row, start_column=2, end_row=second_header_row, end_column=2)

    col = 3

    for schedule in schedules:
        if schedule.has_lesson and schedule.tasks:
            for task in schedule.tasks:
                ws.cell(row=second_header_row, column=col).value = task.name
                ws.cell(row=second_header_row, column=col).fill = header_fill
                ws.cell(row=second_header_row, column=col).font = header_font
                ws.cell(row=second_header_row, column=col).alignment = center_alignment
                ws.cell(row=second_header_row, column=col).border = thin_border
                col += 1
        elif schedule.has_lesson:
            col += 1

    ws.row_dimensions[first_header_row].height = 40
    ws.row_dimensions[second_header_row].height = 30

    row = second_header_row + 1

    for idx, student in enumerate(students, start=1):
        col = 1
        ws.cell(row=row, column=col).value = idx
        ws.cell(row=row, column=col).alignment = center_alignment
        ws.cell(row=row, column=col).border = thin_border
        col += 1

        ws.cell(row=row,
                column=col).value = f"{student.user.last_name} {student.user.first_name} {student.user.patronymic}"
        ws.cell(row=row, column=col).alignment = left_alignment
        ws.cell(row=row, column=col).border = thin_border
        col += 1

        for schedule in schedules:
            if schedule.has_lesson and schedule.tasks:
                for task in schedule.tasks:
                    grade = Grade.objects.filter(task=task, student=student).first()
                    grade_value = grade.value if grade else ''
                    ws.cell(row=row, column=col).value = grade_value
                    ws.cell(row=row, column=col).alignment = center_alignment
                    ws.cell(row=row, column=col).border = thin_border

                    if grade_value == 1 or grade_value == 2:
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="ffcccc", end_color="ffcccc",
                                                                        fill_type="solid")
                    elif grade_value == 3:
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="ffffcc", end_color="ffffcc",
                                                                        fill_type="solid")
                    elif grade_value == 4 or grade_value == 5:
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="ccffcc", end_color="ccffcc",
                                                                        fill_type="solid")

                    col += 1
            elif schedule.has_lesson:
                ws.cell(row=row, column=col).value = ''
                ws.cell(row=row, column=col).alignment = center_alignment
                ws.cell(row=row, column=col).border = thin_border
                col += 1

        row += 1

    for col_num in range(1, ws.max_column + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15

    current_date = datetime.now().strftime('%d.%m.%Y')
    filename = f"{discipline.plan.name}_{discipline.group.name}_{current_date}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    wb.save(response)
    return response